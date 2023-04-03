import io

import openpyxl

# fmt: off

duplicate_user_import_filedata = {
    'Users': [
        ['Title', 'First Name', 'Last Name', 'Email'],
        ['', 'Lucilia', 'Manilium', 'lucilia.manilium@institution.example.com'],
        ['', 'Bastius', 'Quid', 'bastius.quid@external.example.com'],
        ['', 'Bastius', 'Quid', 'bastius.quid@external.example.com']
    ]
}

mismatching_user_import_filedata = {
    'Users': [
        ['Title', 'First Name', 'Last Name', 'Email'],
        ['', 'Lucilia', 'Manilium', 'lucilia.manilium@institution.example.com'],
        ['', 'Bastius', 'Quid', 'bastius.quid@external.example.com'],
        ['Dr.', 'Bastius', 'Quid', 'bastius.quid@external.example.com']
    ]
}

numerical_data_in_user_data_filedata = {
    'Users': [
        ['Title', 'First Name', 'Last Name', 'Email'],
        ['', 'Lucilia', 'Manilium', 'lucilia.manilium@institution.example.com'],
        [3.14, 'Bastius', 'Quid', 'bastius.quid@external.example.com'],
        ['', 'Gustav', 42, 'gustav42@external.example.com']
    ]
}

valid_user_import_filedata = {
    'Skipped Sheet': [["Just", "Some", "Comments"]],
    'Users': [
        ['Title', 'First name', 'Last name', 'Email'],
        ['', 'Lucilia', 'Manilium', 'lucilia.manilium@institution.example.com'],
        ['', 'Bastius', 'Quid', 'bastius.quid@external.example.com'],
    ]
}

missing_values_user_import_filedata = {
    'Sheet 1': [
        ['Title', 'First Name', 'Last Name', 'Email'],
        ['', '', 'Manilium', 'missing.firstname@institution.example.com'],
        ['', 'Lucilia', '', 'missing.lastname@institution.example.com'],
        ['', 'Bastius', 'Quid', ''],
    ]
}

invalid_enrollment_data_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Master', 'Quid', 'Basti', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Herbringen', 'Bring', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Machen', 'Do', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'maybe', 'Verhandeln', 'Deal', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kaufen', 'Buy', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Master', 'Synephebos', 'Diam', '', 'Seminar', 'no', 'Zerplatzen', 'Burst', 'Dr.', 'Sadipscing', 'Elitr', '234@external.example.com'],
        ['Diploma', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Vorlesung', 'yes', 'Schneiden', 'Cut', 'Dr.', 'Sic', 'Graecis', '890@external.example.com'],
        ['Bachelor', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kommen', 'Come', 'Prof. Dr.', 'Takimata', 'Labore', '678@internal.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kosten', 'Cost', 'Dr.', 'Aliquyam', 'Sanctus', ''],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Praktikum', 'no', 'Wählen', 'Choose', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Vorlesung', 'no', 'Schlagen', 'Beat', 'Prof. Dr.', 'Amet', 'Consetetur', '123@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Herbringen', 'Bring', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Zerplatzen', 'Burst', 'Dr.', 'Sadipscing', 'Elitr', '234@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Vorlesung', 'yes', 'Fangen', 'Catch', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Zerbrechen', 'Break', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Kaufen', 'Bought', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Diploma', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Vorlesung', 'yes', 'Schneiden', 'Cut', 'Dr.', 'Sic', 'Graecis', '890@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Kosten', 'Cost', 'Dr.', 'Aliquyam', 'Sanctus', '567@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Praktikum', 'no', 'Wählen', 'Choose', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Zerbrechen', 'Break', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com']
    ],
    'BA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'yes', 'Schütteln', 'Shake', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Singen', 'Sing', 'Dr.', 'Praeterea', 'Eadamque', '345@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Sinken', 'Sink', 'Dr.', 'Itaque', 'Ferdi', '789@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Sitzen', 'Sit', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Schließen', 'Shut', 'Prof. Dr.', 'Itch', 'Probabo', '234@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Zeigen', 'Show', 'Dr.', 'Sed', 'Tam', '456@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Scheinen', 'Shine', 'Prof. Dr.', 'Multi', 'Augendas', '567@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'yes', 'Schütteln', 'Shake', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Seminar', 'yes', 'Stehlen', 'Steal', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'yes', 'Sprechen', 'Speak', 'Prof.-Dr.', 'Honoris', 'Invitat', '111@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Schlafen', 'Sleep', 'Prof. Dr. ', 'Takimata', 'Labore', '678@external.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Zeigen', 'Show', 'Dr.', 'Sed', 'Tam', '456@institution.example.com']
    ]
}

test_enrollment_data_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Herbringen', 'Bring', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Machen', 'Do', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Verhandeln', 'Deal', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kaufen', 'Buy', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Zerplatzen', 'Burst', 'Dr.', 'Sadipscing', 'Elitr', '234@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Vorlesung', 'yes', 'Schneiden', 'Cut', 'Dr.', 'Sic', 'Graecis', '890@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kommen', 'Come', 'Prof. Dr.', 'Takimata', 'Labore', '678@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Kosten', 'Cost', 'Dr.', 'Aliquyam', 'Sanctus', '567@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Vorlesung', 'no', 'Wählen', 'Choose', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Vorlesung', 'no', 'Schlagen', 'Beat', 'Prof. Dr.', 'Amet', 'Consetetur', '123@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Herbringen', 'Bring', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Zerplatzen', 'Burst', 'Dr.', 'Sadipscing', 'Elitr', '234@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Vorlesung', 'yes', 'Fangen', 'Catch', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Master', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Zerbrechen', 'Break', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Kaufen', 'Buy', 'Dr.', 'Romano', 'Electram', '111@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Vorlesung', 'yes', 'Schneiden', 'Cut', 'Dr.', 'Sic', 'Graecis', '890@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Kosten', 'Cost', 'Dr.', 'Aliquyam', 'Sanctus', '567@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Vorlesung', 'no', 'Wählen', 'Choose', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com'],
        ['Master', 'Menandri', 'Latinas', 'latinas.menandri@institution.example.com', 'Seminar', 'no', 'Zerbrechen', 'Break', 'Prof. Dr.', 'Dolor', 'Sit', 'asd@external.example.com']
    ],
    'BA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'yes', 'Schütteln', 'Shake', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Singen', 'Sing', 'Dr.', 'Praeterea', 'Eadamque', '345@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Sinken', 'Sink', 'Dr.', 'Itaque', 'Ferdi', '789@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Sitzen', 'Sit', 'Prof. Dr.', 'Tempor', 'Invidunt', '789@external.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Schließen', 'Shut', 'Prof. Dr.', 'Itch', 'Probabo', '234@institution.example.com'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'Vorlesung', 'no', 'Zeigen', 'Show', 'Dr.', 'Sed', 'Tam', '456@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Scheinen', 'Shine', 'Prof. Dr.', 'Multi', 'Augendas', '567@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'yes', 'Schütteln', 'Shake', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Seminar', 'yes', 'Stehlen', 'Steal', 'Dr.', 'Nonumy', 'Eirmod', '456@external.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'yes', 'Sprechen', 'Speak', 'Prof.-Dr.', 'Honoris', 'Invitat', '111@institution.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Schlafen', 'Sleep', 'Prof. Dr. ', 'Takimata', 'Labore', '678@external.example.com'],
        ['Bachelor', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'Vorlesung', 'no', 'Zeigen', 'Show', 'Dr.', 'Sed', 'Tam', '456@institution.example.com']
    ]
}

test_enrollment_data_empty_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
    ],
    'BA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
    ]
}

test_enrollment_data_consecutive_and_trailing_spaces_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Haeuser Bauen', 'Build     houses ', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Master', 'Synephebos', 'Diam', 'diam.synephebos@institution.example.com', 'Seminar', 'no', 'Haeuser    Bauen  ', 'Build houses', '', 'Sed', 'Diam', '345@external.example.com']
    ]
}

test_enrollment_data_degree_merge_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Bachelor', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com']
    ]
}

test_enrollment_data_error_merge_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Grandmaster', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'jaminar', 'probably not', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['Beginner', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'jaminar', 'probably not', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
    ],
}

test_enrollment_data_import_names_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'S', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com'],
        ['M. Sc.', 'Lorem', 'Ipsum', 'ipsum.lorem@institution.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.example.com']
    ],
    'BA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@institution.example.com', 'VORlesung', 'yes', 'Rechtschreibung', 'Spelling', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['B. Sc.', 'Metrodorus', 'Torquate', 'torquate.metrodorus@institution.example.com', 'V', 'yes', 'Rechtschreibung', 'Spelling', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com']
    ]
}

test_unknown_degree_error_filedata = {
    'Sheet 1': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['bachelor', 'Doe', 'John', 'john.doe@institution.example.com', 'Vorlesung', 'yes', 'Neovim kompilieren', 'compiling Neovim', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
        ['beginner', 'Roe', 'Jane', 'jane.roe@institution.example.com', 'Vorlesung', 'yes', 'Neovim kompilieren', 'compiling Neovim', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@institution.example.com'],
    ],
}

valid_user_courses_import_filedata = {
    'MA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Master', 'Quid', 'Bastius', 'bastius.quid@external.example.com', 'Seminar', 'no', 'Bauen', 'Build', '', 'Sed', 'Diam', '345@external.institution.com']
    ],
    'BA Belegungen': [
        ['Degree', 'Student last name', 'Student first name', 'Student email address', 'Course kind', 'Course is graded', 'Course name (de)', 'Course name (en)', 'Responsible title', 'Responsible last name', 'Responsible first name', 'Responsible email address'],
        ['Bachelor', 'Manilium', 'Lucilia', 'lucilia.manilium@student.institution.com', 'Vorlesung', 'no', 'Schütteln', 'Shake', 'Prof. Dr.', 'Prorsus', 'Christoph', '123@external.com']
    ]
}

random_file_content = "Hallo Welt\n".encode()


wrong_column_count_excel_data = {
    'Sheet 1': [
        ["Only", "Three", "Columns"],
        ["Some", "Content", "Here"],
    ]
}
# fmt: on


def create_memory_excel_file(data):
    memory_excel_file = io.BytesIO()
    workbook = openpyxl.Workbook()
    for sheet_name, sheet_data in data.items():
        sheet = workbook.create_sheet(sheet_name)
        for row_num, row_data in enumerate(sheet_data, 1):
            for column_num, cell_data in enumerate(row_data, 1):
                # openpyxl rows start at 1
                sheet.cell(row=row_num, column=column_num).value = cell_data
    workbook.save(memory_excel_file)
    return memory_excel_file.getvalue()
