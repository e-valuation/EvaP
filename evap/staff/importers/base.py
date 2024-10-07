import itertools
from abc import ABC, abstractmethod
from collections import Counter, namedtuple
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from enum import Enum
from io import BytesIO
from typing import Any

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.utils.translation import gettext as _
from django.utils.translation import gettext_lazy, ngettext


@dataclass
class ImporterLogEntry:
    class Level(Enum):
        ERROR = 2
        WARNING = 1
        SUCCESS = 0

    class Category(Enum):
        _CATEGORY_TUPLE = namedtuple("_CATEGORY_TUPLE", ["id", "display_name", "order"])

        # used to make sure result messages come first
        RESULT = _CATEGORY_TUPLE("result", gettext_lazy("Result"), -1)
        GENERAL = _CATEGORY_TUPLE("general", gettext_lazy("General"), 0)

        SCHEMA = _CATEGORY_TUPLE("schema", gettext_lazy("Incorrect Excel format"), 1)

        PROGRAM_MISSING = _CATEGORY_TUPLE("missing_program", gettext_lazy("Missing programs"), 2)
        COURSE_TYPE_MISSING = _CATEGORY_TUPLE("missing_course_type", gettext_lazy("Missing course types"), 3)
        COURSE = _CATEGORY_TUPLE("course", gettext_lazy("Course issues"), 4)
        IS_GRADED = _CATEGORY_TUPLE("is_graded", gettext_lazy("Invalid values"), 5)

        USER = _CATEGORY_TUPLE("user", gettext_lazy("Invalid user data"), 6)

        NAME = _CATEGORY_TUPLE("name", gettext_lazy("Name mismatches"), 7)
        INACTIVE = _CATEGORY_TUPLE("inactive", gettext_lazy("Inactive users"), 8)
        DUPL = _CATEGORY_TUPLE("duplicate", gettext_lazy("Possible duplicates"), 9)
        EXISTS = _CATEGORY_TUPLE("existing", gettext_lazy("Existing courses"), 10)
        IGNORED = _CATEGORY_TUPLE("ignored", gettext_lazy("Ignored duplicates"), 11)
        ALREADY_PARTICIPATING = _CATEGORY_TUPLE("already_participating", gettext_lazy("Existing participants"), 12)

        PROGRAM = _CATEGORY_TUPLE("program", gettext_lazy("Program mismatches"), 13)
        TOO_MANY_ENROLLMENTS = _CATEGORY_TUPLE(
            "too_many_enrollments", gettext_lazy("Unusually high number of enrollments"), 14
        )

        SIMILAR_COURSE_NAMES = _CATEGORY_TUPLE("similar_course_names", gettext_lazy("Similar course names"), 15)

    level: Level
    category: Category
    message: str


class ImporterLog:
    """Just a fancy wrapper around a collection of messages with some utility functions"""

    def __init__(self) -> None:
        self.messages: list[ImporterLogEntry] = []

    def __repr__(self) -> str:
        return f"{type(self).__name__}({self.messages})"  # pragma: no cover

    def _messages_with_level_sorted_by_category(self, level: ImporterLogEntry.Level) -> list[ImporterLogEntry]:
        return sorted(
            (msg for msg in self.messages if msg.level == level),
            key=lambda msg: (msg.category.value.order, msg.category.value.id),
        )

    def _messages_with_level_by_category(
        self, level: ImporterLogEntry.Level
    ) -> dict[ImporterLogEntry.Category, list[ImporterLogEntry]]:
        sorted_messages = self._messages_with_level_sorted_by_category(level)
        grouped_messages = itertools.groupby(sorted_messages, lambda msg: msg.category)
        return {category: list(messages) for category, messages in grouped_messages}

    def add_message(self, message: ImporterLogEntry):
        self.messages.append(message)

    def has_errors(self) -> bool:
        return any(msg.level == ImporterLogEntry.Level.ERROR for msg in self.messages)

    def raise_if_has_errors(self) -> None:
        if self.has_errors():
            raise ImporterError(message="")

    def success_messages(self) -> list[ImporterLogEntry]:
        return self._messages_with_level_sorted_by_category(ImporterLogEntry.Level.SUCCESS)

    def warnings_by_category(self) -> dict[ImporterLogEntry.Category, list[ImporterLogEntry]]:
        return self._messages_with_level_by_category(ImporterLogEntry.Level.WARNING)

    def errors_by_category(self) -> dict[ImporterLogEntry.Category, list[ImporterLogEntry]]:
        return self._messages_with_level_by_category(ImporterLogEntry.Level.ERROR)

    def forward_messages_to_django(self, request) -> None:
        method_by_level = {
            ImporterLogEntry.Level.SUCCESS: messages.success,
            ImporterLogEntry.Level.WARNING: messages.warning,
            ImporterLogEntry.Level.ERROR: messages.error,
        }

        for message in self.messages:
            method_by_level[message.level](request, message.message)

    def add_error(self, message_text, *, category=ImporterLogEntry.Category.GENERAL):
        return self.add_message(ImporterLogEntry(ImporterLogEntry.Level.ERROR, category, message_text))

    def add_warning(self, message_text, *, category=ImporterLogEntry.Category.GENERAL):
        return self.add_message(ImporterLogEntry(ImporterLogEntry.Level.WARNING, category, message_text))

    def add_success(self, message_text, *, category=ImporterLogEntry.Category.GENERAL):
        return self.add_message(ImporterLogEntry(ImporterLogEntry.Level.SUCCESS, category, message_text))


class ImporterError(Exception):
    """Used to abort the import run immediately"""

    def __init__(
        self,
        *args,
        message: str,
        category: ImporterLogEntry.Category = ImporterLogEntry.Category.GENERAL,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)
        self.message = message
        self.category = category

    def as_importer_message(self) -> ImporterLogEntry:
        return ImporterLogEntry(ImporterLogEntry.Level.ERROR, self.category, self.message)


class ConvertExceptionsToMessages:
    """Shared catch-all exception handlers between importers"""

    def __init__(self, importer_log: ImporterLog):
        self.importer_log = importer_log

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_value, traceback) -> bool:
        if isinstance(exc_value, ImporterError):
            # Importers raise these to immediately abort with a message
            if exc_value.message:
                self.importer_log.add_message(exc_value.as_importer_message())

            self.importer_log.add_error(
                _("Errors occurred while parsing the input data. No data was imported."),
                category=ImporterLogEntry.Category.RESULT,
            )
            return True  # do not propagate the error

        if isinstance(exc_value, Exception):
            # Something unexpected happened.
            self.importer_log.add_error(
                _("Import aborted after exception: '{}'. No data was imported.").format(exc_value)
            )
            return not settings.DEBUG  # only propagate the error if settings.DEBUG is True

        return False


@dataclass(eq=True, frozen=True)
class ExcelFileLocation:
    sheet_name: str
    row_number: int

    def __str__(self) -> str:
        return _('Sheet "{}", row {}').format(self.sheet_name, self.row_number + 1)


class InputRow(ABC):
    # MyPy is currently broken with abstract properties: https://github.com/python/mypy/issues/8996
    column_count: int

    @abstractmethod
    def __init__(self, location: ExcelFileLocation, *cells: Iterable[str]):
        pass

    @classmethod
    def from_cells(cls, location: ExcelFileLocation, cells: Iterable[str]):
        return cls(location, *cells)


class ExcelFileRowMapper:
    """
    Take a excel file and map it to a list of row_cls instances
    """

    def __init__(self, skip_first_n_rows: int, row_cls: type[InputRow], importer_log: ImporterLog):
        self.skip_first_n_rows = skip_first_n_rows
        self.row_cls = row_cls
        self.importer_log = importer_log

    def map(self, file_content: bytes):
        try:
            book = openpyxl.load_workbook(BytesIO(file_content))
        except Exception as e:  # noqa: BLE001
            raise ImporterError(
                message=_("Couldn't read the file. Error: {}").format(e),
                category=ImporterLogEntry.Category.SCHEMA,
            ) from e

        rows = []
        for sheet in book:  # type: ignore[attr-defined]
            if sheet.max_row <= self.skip_first_n_rows:
                continue

            if sheet.max_column != self.row_cls.column_count:
                raise ImporterError(
                    message=_("Wrong number of columns in sheet '{}'. Expected: {}, actual: {}").format(
                        sheet.title, self.row_cls.column_count, sheet.max_column
                    )
                )

            # openpyxl uses 1-based indexing.
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=self.skip_first_n_rows + 1, values_only=True), start=self.skip_first_n_rows
            ):
                location = ExcelFileLocation(sheet.title, row_number)

                if not all(isinstance(cell, str) or cell is None for cell in row):
                    self.importer_log.add_error(
                        _(
                            "{location}: Wrong data type. Please make sure all cells are string types, not numerical."
                        ).format(location=location),
                        category=ImporterLogEntry.Category.SCHEMA,
                    )
                    continue

                raw_cells = [cell if cell is not None else "" for cell in row]
                cells = [" ".join(cell.split()) for cell in raw_cells]

                # expand up to column_count values to prevent errors with empty fields
                cells += [""] * (self.row_cls.column_count - len(cells))

                rows.append(self.row_cls.from_cells(location, cells))

            self.importer_log.raise_if_has_errors()
            self.importer_log.add_success(_("Successfully read sheet '%s'.") % sheet.title)

        self.importer_log.add_success(_("Successfully read Excel file."))

        return rows


class FirstLocationAndCountTracker:
    """Track locations by a key to only give a single aggregated message with first occurence and count"""

    def __init__(self, *args, **kwargs) -> None:
        self.first_location_by_key: dict[Any, ExcelFileLocation] = {}
        self.location_count_by_key: Counter = Counter()

    def add_location_for_key(self, location: ExcelFileLocation, key: Any):
        self.first_location_by_key.setdefault(key, location)
        self.location_count_by_key.update([key])

    def aggregated_keys_and_location_strings(self) -> Iterator[tuple[Any, str]]:
        for key, first_location in self.first_location_by_key.items():
            count = self.location_count_by_key[key]

            if count == 1:
                location_string = str(first_location)
            else:
                location_string = ngettext(
                    "{location} and {count} other place", "{location} and {count} other places", count - 1
                ).format(
                    location=first_location,
                    count=count - 1,
                )

            yield key, location_string

    def keys(self) -> Iterable[Any]:
        return self.first_location_by_key.keys()


class Checker:
    def __init__(self, test_run: bool, importer_log: ImporterLog):
        self.test_run = test_run
        self.importer_log = importer_log

    def finalize(self) -> None:
        # can be overriden by implementations to generate aggregated messages
        pass


class RowCheckerMixin(ABC):
    @abstractmethod
    def finalize(self) -> None:
        pass  # pragma: no cover

    @abstractmethod
    def check_row(self, row) -> None:
        pass  # pragma: no cover

    def check_rows(self, rows) -> None:
        for row in rows:
            self.check_row(row)
        self.finalize()
