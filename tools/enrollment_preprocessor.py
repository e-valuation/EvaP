#!/usr/bin/env python3

import csv
import sys
from argparse import ArgumentParser
from collections import defaultdict
from collections.abc import Iterator
from datetime import datetime
from io import BytesIO
from itertools import chain
from pathlib import Path
from typing import NamedTuple, TextIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell

_stdout = sys.stdout


class User(NamedTuple):
    title: str
    last_name: str
    first_name: str
    email: str

    def full_name(self) -> str:
        prefix = ""
        if self.title:
            prefix = f"{self.title} "
        return f"{prefix}{self.first_name} {self.last_name}, {self.email}"


class UserCells(NamedTuple):
    title: Cell | None
    last_name: Cell
    first_name: Cell
    email: Cell

    def _clean(self) -> Iterator[str]:
        for field in iter(self):
            if not field or not field.value:
                yield ""
                continue
            field.value = str(field.value).strip()
            yield field.value

    def clean_user(self):
        return User(*self._clean())

    def update_from(self, user: User) -> bool:
        changed = False
        for cell, field_value in zip(self, [user.title, user.last_name, user.first_name, user.email], strict=True):
            if cell is not None and cell.value != field_value:
                changed = True
                cell.value = field_value
        return changed


def user_from_row(row: tuple[Cell, ...]):
    return [
        UserCells(None, *row[:3]).clean_user(),
        UserCells(*row[7:]).clean_user(),
    ]


def make_bold(text: str) -> str:
    return f"\033[1m{text}\033[0m"


def csv_users_by_email(user_data: TextIO) -> dict[str, User]:
    users = {}
    reader = csv.reader(user_data, delimiter=";", lineterminator="\n")
    next(reader)  # skip header
    for row in reader:
        user = User(*row)
        users[user.email] = user
    return users


def get_user_decisions(database_users: dict[str, User], workbook: Workbook) -> dict[str, User]:
    conflicts_by_field: dict[str, set[User]] = defaultdict(set)
    for sheet in workbook.worksheets:  # parse enrollment data and group conflicts
        for imported in chain.from_iterable(map(user_from_row, sheet.iter_rows(min_row=2, min_col=2))):
            if not imported.email:
                continue
            existing = database_users.setdefault(imported.email, imported)
            for field in ["title", "last_name", "first_name"]:
                field_value = getattr(imported, field)
                if field_value is not None and getattr(existing, field) != getattr(imported, field):
                    conflicts_by_field[field].add(imported)
    # ask user for decision
    for field, conflicts in conflicts_by_field.items():
        _stdout.write(f"{field.capitalize()}\n")
        _stdout.write("---------\n")
        for imported in sorted(conflicts):
            existing = database_users[imported.email]

            if getattr(existing, field) == getattr(imported, field):
                continue

            _stdout.write(f"existing: '{make_bold(getattr(existing, field))}' ({existing.full_name()})\n")
            _stdout.write(f"imported: '{make_bold(getattr(imported, field))}' ({imported.full_name()})\n")

            decision = ""
            while decision not in ("e", "i"):
                decision = input("Which one should be used? (e/i):\n")
            choice = existing if decision == "e" else imported

            database_users[choice.email] = database_users[choice.email]._replace(**{field: getattr(choice, field)})
        _stdout.write("\n")
    return database_users


def run_preprocessor(enrollment_data: Path | BytesIO, user_data: TextIO) -> BytesIO | None:
    workbook = load_workbook(enrollment_data)

    correct_user_data_by_email = get_user_decisions(
        csv_users_by_email(user_data),
        workbook,
    )

    # apply decisions
    changed = False
    for sheet in workbook.worksheets:
        for wb_row in sheet.iter_rows(min_row=2, min_col=2):
            if wb_row[2] and wb_row[2].value:
                changed |= UserCells(None, *wb_row[:3]).update_from(correct_user_data_by_email[wb_row[2].value.strip()])

            if wb_row[-1] and wb_row[-1].value:
                changed |= UserCells(*wb_row[7:]).update_from(correct_user_data_by_email[wb_row[-1].value.strip()])

    if not changed:
        return None
    wb_out = BytesIO()
    workbook.save(wb_out)
    wb_out.seek(0)
    return wb_out


if __name__ == "__main__":  # pragma: nocover
    parser = ArgumentParser(description="Commandline tool to preprocess enrollment xlsx files.")
    parser.add_argument("user_data", help="Path to a csv file containing an export of all existing users.")
    parser.add_argument("enrollment_data", help="Path to the enrollment data in xlsx format for import.")
    ns = parser.parse_args()

    target = Path(ns.enrollment_data)

    with open(ns.user_data, encoding="utf-8") as csvfile:
        wb = run_preprocessor(target, csvfile)
        if wb is None:
            _stdout.write("Done! No changes to the excel file were necessary!\n")
            sys.exit()
    with target.with_stem(f"{target.stem}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}").open("wb") as out:
        _stdout.write("Done! All conflicts are resolved in a new file!\n")
        out.write(wb.read())
